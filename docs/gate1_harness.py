#!/usr/bin/env python3
"""
AltifyOS Gate 1 harness — v2, ratification-review protocol.

WRITES NOTHING ANYWHERE except its own local output files.
No Salesforce client is imported. No Apex. No org writes. Runs entirely outside
the org: read a transcript from disk, call a model, string-check every cited
span, re-adjudicate blind, then hand the result to Matthew for ratification.

THE PROTOCOL — ruled 29 July 2026, replacing pre-labeled scoring
----------------------------------------------------------------
Per transcript:
  1. extract       — the machine commits verdicts, each carrying 1-5 verbatim
                     spans (multi-span citation, per the Architecture ruling).
  2. readjudicate  — a blind second reader judges each span SET against its
                     claim, seeing nothing else.
  3. review        — produces a spreadsheet for Matthew: every committed
                     verdict with its evidence and the blind read, for
                     Agree / Disagree + reason; then every unanswered
                     proposition, for a miss sweep: Confirmed empty /
                     Missed - minor / Missed - material.
  4. grade         — reads the completed sheet back and applies the bar.

THE BAR — a transcript passes when ALL of:
  - zero fabricated or turn-crossing spans (mechanical, from extraction)
  - ratification rate >= 0.90 (Matthew agrees with >= 9 in 10 committed verdicts)
  - zero material misses (no case where evidence a competent seller would
    have flagged went unanswered)
Any fabrication or any material miss fails the transcript outright.
Matthew's ratified sheet is the ground truth; his disagree-reasons feed the
charter, and his agree/disagree is a manual rehearsal of the product's own
ratification gate. Planned, not built: 10x high-end-model ensemble runs as a
later variance probe.

TRANSCRIPT PRECONDITION — machine annotations (Einstein topic labels and any
other inline non-spoken text) MUST be stripped upstream, before the transcript
JSON is written. A model citing an annotation would pass a naive span check.
The harness cannot detect them reliably; the retrieval step owns the strip and
must record it.

Usage
-----
  export ANTHROPIC_API_KEY=...
  python gate1_harness.py extract      --transcript t2.json --mode whole
  python gate1_harness.py extract      --transcript t2.json --mode chunked --chunk-turns 40
  python gate1_harness.py readjudicate --run runs/<run_id>
  python gate1_harness.py review       --run runs/<run_id>
  ... Matthew fills review_sheet.xlsx ...
  python gate1_harness.py grade        --run runs/<run_id>

Transcript input format (JSON) — produced by whatever read-only retrieval you
use; the harness does not fetch it:
  {"turns": [{"speaker": "Full Name", "email": "a@b.com", "start_s": 12.0,
              "text": "..."}, ...]}
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import unicodedata
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

# --------------------------------------------------------------------------
# Configuration that is a JUDGMENT, not a fact. Named here so it is visible.
# --------------------------------------------------------------------------

MODEL = os.environ.get("GATE1_MODEL", "claude-opus-5")

# Ruling: speaker role derives from EMAIL DOMAIN, never from platform record
# resolution. Specimen hazard: Patrick Morrissey holds an altify.com address but
# resolves to a Lead in the org. Domain is the observable; resolution is not.
INTERNAL_DOMAINS = {"altify.com", "upland.com", "uplandsoftware.com"}

# Multi-span citation — graduated ruling, Architecture. A verdict carries one
# to five verbatim spans; each is separately checked and must sit inside a
# single speaker turn; the SET carries the verdict and fails if any member fails.
MAX_SPANS = 5

# The pass bar — Matthew's ruling, 29 July 2026. Changing a number here is a
# decision, not a tweak; record it in the documents.
GATE1_BAR = {
    "max_fabricated_or_crossing_spans": 0,
    "min_ratification_rate": 0.90,
    "max_material_misses": 0,
}

# Span checking. Both strictnesses are computed and BOTH are reported.
#   exact       — the cited string appears verbatim in the transcript, byte for byte
#   normalized  — appears after unicode NFKC, lowercasing, quote/dash folding,
#                 and whitespace collapse
# A span passing normalized but failing exact is a transcription-fidelity
# artefact, not necessarily a fabrication. A span failing both is fabricated
# and the run's most important number.

TRUE, FALSE, ND = "true", "false", "not discussed"
VERDICTS = (TRUE, FALSE, ND)

# --------------------------------------------------------------------------
# Output schema — three states, structurally distinguished
# --------------------------------------------------------------------------
# null    : the model returned nothing for this proposition (a harness/parse
#           failure, NOT an answer). Never conflated with abstention.
# cited   : verdict true|false WITH 1-5 verbatim spans. The only establishing state.
# abstain : verdict "not discussed", with a required reason:
#             model_missed  — the model believes the evidence may be present but
#                             it could not locate or commit to a span
#             nobody_said   — the model asserts the transcript does not contain it
#           The split is the whole point: model_missed is our failure,
#           nobody_said is the world's.

ABSTAIN_REASONS = ("model_missed", "nobody_said")

RESPONSE_SCHEMA = {
    "type": "object",
    "properties": {
        "establishments": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "proposition_id": {"type": "string"},
                    "state": {"type": "string", "enum": ["cited", "abstain"]},
                    "verdict": {"type": "string", "enum": list(VERDICTS)},
                    "spans": {
                        "type": ["array", "null"],
                        "minItems": 1,
                        "maxItems": MAX_SPANS,
                        "items": {
                            "type": "object",
                            "properties": {
                                "quote": {
                                    "type": "string",
                                    "description": "Verbatim, contiguous substring of ONE "
                                                   "speaker turn. Character for character.",
                                },
                                "speaker": {"type": "string"},
                            },
                            "required": ["quote", "speaker"],
                            "additionalProperties": False,
                        },
                        "description": "1-5 spans. Required when state=cited. "
                                       "Must be null when state=abstain.",
                    },
                    "abstain_reason": {"type": ["string", "null"], "enum": list(ABSTAIN_REASONS) + [None]},
                    "confidence": {"type": "string", "enum": ["high", "medium", "low"]},
                },
                "required": ["proposition_id", "state", "verdict", "spans",
                             "abstain_reason", "confidence"],
                "additionalProperties": False,
            },
        }
    },
    "required": ["establishments"],
    "additionalProperties": False,
}

# --------------------------------------------------------------------------
# The charter prompt
# --------------------------------------------------------------------------

CHARTER = """You are an evidence extractor for a methodology-enforcement system. You VERIFY. You never PREDICT.

Your only job: for each numbered proposition below, decide whether this transcript establishes it true, establishes it false, or does not discuss it — and, when it establishes either, quote the exact words that did so.

BINDING RULES

1. Every true and every false MUST carry between one and five `spans`. Each span's `quote` is a verbatim, contiguous substring of a SINGLE speaker turn in the transcript you were given. Copy it character for character. Do not clean up grammar, fix punctuation, remove filler, join across a speaker change, or elide with "...". If you cannot produce such spans, you do not have a verdict — abstain.

2. The span SET must do the establishing on its own. One decisive sentence needs one span. A fact established by accumulation — several acts or statements that only together carry the claim — needs a span per act. If a reader seeing only your spans would not conclude the verdict, they are the wrong spans. Do not cite context that merely surrounds the fact, and do not stack hedged fragments to imply a commitment none of them makes.

3. Wrong-and-confident is the single unacceptable failure. Abstaining costs almost nothing. Inferring, generalising, or reading a plausible implication as a statement costs everything. When the transcript nearly says it, that is a `not discussed`.

4. Abstention is not one thing. Every abstention needs a reason:
   - `model_missed` — you think the evidence may well be in here and you could not pin it to spans you would defend.
   - `nobody_said` — you assert the transcript does not contain this. Nobody raised it.
   Choose deliberately. This split is measured.

5. Speaker requirements bind. Where a proposition names a required speaker, only that side can establish it.
   - Internal (seller side) speakers are identified by EMAIL DOMAIN: {internal_domains}.
   - Everyone else is external (customer side), regardless of what any system says about them.
   - A seller asserting a customer fact does not establish it. A seller repeating what a customer told them does not establish it either — the customer must have said it here.
   - A customer relaying a third party's position establishes only that the customer said it, not the third party's position.
   - Every span names its own speaker, exactly as that speaker is named in the transcript.

6. Answer every proposition exactly once. Return the full list. Never invent proposition ids.

7. Return only the JSON object. No prose before or after.

PARTICIPANTS
{participants}

PROPOSITIONS
{propositions}

TRANSCRIPT
{transcript}

Return JSON matching this schema:
{schema}
"""

READJUDICATE = """You are a second, independent adjudicator. You have not seen any prior verdict and you must not try to guess one.

For each item you are given a proposition and a SET of one or more quoted spans, each said by its named speaker. Decide only this: does that set of spans, taken together and on its own, establish the proposition true, establish it false, or neither?

Rules:
- Judge the spans alone. You have no other context and must not supply any.
- The set fails if it depends on any member that does not hold up. "Neither" includes: the spans are about a related but different thing; they are suggestive but not assertive; a required speaker is the wrong side; the set is hedged fragments stacked to imply a commitment none of them makes; the claim would need context beyond these spans.
- Where the proposition is a compound (two or more clauses joined by "and"), every clause must be carried by the set. Say which clause fails if one does.
- Speaker sides: internal (seller) domains are {internal_domains}; anyone else is external (customer).

ITEMS
{items}

Return only JSON:
{{"adjudications": [{{"proposition_id": "...", "verdict": "true|false|not discussed", "reason": "one short clause"}}]}}
"""

# --------------------------------------------------------------------------
# Normalization + span checking
# --------------------------------------------------------------------------

_QUOTES = {
    "\u2018": "'", "\u2019": "'", "\u201a": "'", "\u201b": "'",
    "\u201c": '"', "\u201d": '"', "\u201e": '"', "\u201f": '"',
    "\u2013": "-", "\u2014": "-", "\u2212": "-", "\u2026": "...",
    "\u00a0": " ", "\u200b": "",
}


def fold(s: str) -> str:
    """NFKC, quote/dash folding, lowercase, whitespace collapse."""
    s = unicodedata.normalize("NFKC", s)
    for a, b in _QUOTES.items():
        s = s.replace(a, b)
    s = s.lower()
    s = re.sub(r"\s+", " ", s)
    return s.strip()


@dataclass
class SpanResult:
    quote: str
    claimed_speaker: str | None
    exact: bool
    normalized: bool
    turn_index: int | None
    speaker: str | None
    speaker_side: str | None
    char_offset: int | None
    length: int
    crosses_turn: bool
    speaker_mismatch: bool

    @property
    def fabricated(self) -> bool:
        return not self.exact and not self.normalized


class Transcript:
    def __init__(self, turns: list[dict[str, Any]]):
        # Speaker side is derived HERE, on every construction path, so a
        # directly-built or chunked Transcript can never carry an unresolved
        # side. Domain is the observable; platform record resolution is not.
        for t in turns:
            t.setdefault("email", "")
            t["speaker_side"] = side_of(t.get("email", ""))
        self.turns = turns
        self.full = "\n".join(t.get("text", "") for t in turns)
        self.full_folded = fold(self.full)
        self._turn_folded = [fold(t.get("text", "")) for t in turns]

    @classmethod
    def load(cls, path: Path) -> "Transcript":
        data = json.loads(path.read_text(encoding="utf-8"))
        turns = data["turns"] if isinstance(data, dict) else data
        return cls(turns)

    def check(self, quote: str | None, claimed_speaker: str | None = None) -> SpanResult | None:
        if not quote:
            return None
        exact = quote in self.full
        q = fold(quote)
        norm = bool(q) and q in self.full_folded
        turn_i = None
        speaker = None
        side = None
        offset = None
        crosses = False
        for i, tf in enumerate(self._turn_folded):
            if q and q in tf:
                turn_i = i
                speaker = self.turns[i].get("speaker")
                side = self.turns[i].get("speaker_side")
                offset = tf.find(q)
                break
        if norm and turn_i is None:
            # matches the concatenated transcript but no single turn: the model
            # stitched across a speaker change, which rule 1 forbids.
            crosses = True
        mismatch = bool(
            claimed_speaker and speaker
            and fold(claimed_speaker) != fold(speaker)
        )
        return SpanResult(quote, claimed_speaker, exact, norm, turn_i, speaker,
                          side, offset, len(quote), crosses, mismatch)


def side_of(email: str) -> str | None:
    if not email or "@" not in email:
        return None
    return "internal" if email.split("@")[-1].strip().lower() in INTERNAL_DOMAINS else "external"


# --------------------------------------------------------------------------
# Model call
# --------------------------------------------------------------------------

@dataclass
class Usage:
    input_tokens: int = 0
    output_tokens: int = 0
    calls: int = 0
    wall_s: float = 0.0

    def add(self, other: "Usage") -> None:
        self.input_tokens += other.input_tokens
        self.output_tokens += other.output_tokens
        self.calls += other.calls
        self.wall_s += other.wall_s


def call_model(prompt: str, max_tokens: int = 16000) -> tuple[str, Usage]:
    try:
        from anthropic import Anthropic
    except ImportError:
        sys.exit("pip install anthropic")
    client = Anthropic()
    t0 = time.time()
    resp = client.messages.create(
        model=MODEL,
        max_tokens=max_tokens,
        temperature=0,
        messages=[{"role": "user", "content": prompt}],
    )
    wall = time.time() - t0
    text = "".join(b.text for b in resp.content if getattr(b, "type", "") == "text")
    u = Usage(resp.usage.input_tokens, resp.usage.output_tokens, 1, wall)
    return text, u


def parse_json(text: str) -> dict[str, Any]:
    m = re.search(r"\{.*\}", text, re.S)
    if not m:
        raise ValueError("no JSON object in model response")
    return json.loads(m.group(0))


# --------------------------------------------------------------------------
# Extraction
# --------------------------------------------------------------------------

def render_propositions(props: list[dict[str, Any]]) -> str:
    lines = []
    for i, p in enumerate(props, 1):
        bits = [f"{i}. [{p['id']}] {p['text']}"]
        if p.get("speaker"):
            bits.append(f"   Required speaker: {p['speaker']}")
        lines.append("\n".join(bits))
    return "\n".join(lines)


def render_participants(tr: Transcript) -> str:
    seen: dict[str, str] = {}
    for t in tr.turns:
        key = (t.get("speaker") or "?")
        if key not in seen:
            seen[key] = t.get("email", "") or "(no email)"
    return "\n".join(
        f"- {name} <{email}> — {side_of(email) or 'UNKNOWN SIDE'}"
        for name, email in seen.items()
    )


def chunk_turns(tr: Transcript, size: int, overlap: int = 5) -> list[Transcript]:
    out = []
    i = 0
    while i < len(tr.turns):
        out.append(Transcript(tr.turns[i:i + size]))
        i += max(1, size - overlap)
    return out


def extract(args: argparse.Namespace) -> None:
    aset = json.loads(Path(args.applicable_set).read_text(encoding="utf-8"))
    props = aset["propositions"]
    tr = Transcript.load(Path(args.transcript))

    run_id = f"{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}-{args.mode}"
    run_dir = Path(args.out) / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    usage = Usage()
    raw_by_source: list[dict[str, Any]] = []

    sources = [tr] if args.mode == "whole" else chunk_turns(tr, args.chunk_turns)
    for si, src in enumerate(sources):
        prompt = CHARTER.format(
            internal_domains=", ".join(sorted(INTERNAL_DOMAINS)),
            participants=render_participants(src),
            propositions=render_propositions(props),
            transcript=src.full,
            schema=json.dumps(RESPONSE_SCHEMA, indent=2),
        )
        text, u = call_model(prompt)
        usage.add(u)
        try:
            payload = parse_json(text)
        except Exception as e:  # a parse failure is null, never abstention
            payload = {"establishments": [], "_parse_error": str(e), "_raw": text}
        raw_by_source.append({"source_index": si, "turns": len(src.turns), "payload": payload})

    # merge: for chunked runs, a cited beats an abstention; among cited, first wins
    merged: dict[str, dict[str, Any]] = {}
    for entry in raw_by_source:
        for e in entry["payload"].get("establishments", []):
            pid = e.get("proposition_id")
            if pid not in {p["id"] for p in props}:
                continue  # hallucinated id, dropped and counted below
            cur = merged.get(pid)
            if cur is None or (cur.get("state") != "cited" and e.get("state") == "cited"):
                e = dict(e)
                e["_source_index"] = entry["source_index"]
                merged[pid] = e

    # null = model returned nothing for this proposition
    records = []
    for p in props:
        e = merged.get(p["id"])
        base = {
            "proposition_id": p["id"], "route": p["route"],
            "solicit": p.get("solicit", False),
            "existing_human_answer": p.get("existing_human_answer"),
        }
        if e is None:
            records.append({**base, "state": "null", "verdict": None, "spans": [],
                            "abstain_reason": None, "confidence": None, "demoted": None})
            continue

        raw_spans = e.get("spans") or []
        checked = [tr.check(s.get("quote"), s.get("speaker"))
                   for s in raw_spans if isinstance(s, dict)]
        checked = [c for c in checked if c is not None]

        # a cited state whose span set does not verify is DEMOTED, and the
        # demotion is what the fidelity metric counts. The SET fails if any
        # member fails — per the multi-span ruling.
        state = e.get("state")
        demoted = None
        if state == "cited":
            if not checked:
                state, demoted = "null", "cited_without_spans"
            elif len(checked) > MAX_SPANS:
                state, demoted = "rejected", "more_than_five_spans"
            elif any(c.fabricated for c in checked):
                state, demoted = "rejected", "span_not_in_transcript"
            elif any(c.crosses_turn for c in checked):
                state, demoted = "rejected", "span_crosses_speaker_turn"
            elif any(c.speaker_mismatch for c in checked):
                state, demoted = "rejected", "span_speaker_misattributed"
            elif any(not c.exact for c in checked):
                demoted = "span_normalized_only"

        records.append({
            **base,
            "state": state, "verdict": e.get("verdict"),
            "abstain_reason": e.get("abstain_reason"),
            "confidence": e.get("confidence"), "demoted": demoted,
            "spans": [asdict(c) for c in checked],
        })

    hallucinated = sorted({
        e.get("proposition_id")
        for entry in raw_by_source
        for e in entry["payload"].get("establishments", [])
        if e.get("proposition_id") not in {p["id"] for p in props}
    })

    out = {
        "run_id": run_id,
        "model": MODEL,
        "mode": args.mode,
        "chunk_turns": args.chunk_turns if args.mode == "chunked" else None,
        "sources": len(sources),
        "transcript_path": str(args.transcript),
        "transcript_sha256": hashlib.sha256(tr.full.encode()).hexdigest(),
        "transcript_turns": len(tr.turns),
        "applicable_set_provenance": aset.get("_provenance"),
        "internal_domains": sorted(INTERNAL_DOMAINS),
        "usage": asdict(usage),
        "hallucinated_proposition_ids": hallucinated,
        "records": records,
    }
    (run_dir / "extraction.json").write_text(json.dumps(out, indent=2), encoding="utf-8")
    (run_dir / "raw.json").write_text(json.dumps(raw_by_source, indent=2), encoding="utf-8")
    print(json.dumps({
        "run": str(run_dir),
        "cited": sum(1 for r in records if r["state"] == "cited"),
        "abstain": sum(1 for r in records if r["state"] == "abstain"),
        "rejected": sum(1 for r in records if r["state"] == "rejected"),
        "null": sum(1 for r in records if r["state"] == "null"),
        "spans_total": sum(len(r["spans"]) for r in records),
        "normalized_only": sum(1 for r in records if r.get("demoted") == "span_normalized_only"),
        "hallucinated_ids": hallucinated,
        "input_tokens": usage.input_tokens,
        "output_tokens": usage.output_tokens,
        "wall_s": round(usage.wall_s, 1),
    }, indent=2))


# --------------------------------------------------------------------------
# Blind re-adjudication
# --------------------------------------------------------------------------

def readjudicate(args: argparse.Namespace) -> None:
    run_dir = Path(args.run)
    ex = json.loads((run_dir / "extraction.json").read_text(encoding="utf-8"))
    aset = json.loads(Path(args.applicable_set).read_text(encoding="utf-8"))
    by_id = {p["id"]: p for p in aset["propositions"]}

    items = [r for r in ex["records"] if r["state"] == "cited" and r.get("spans")]
    if not items:
        print(json.dumps({"note": "no cited spans to re-adjudicate"}))
        return

    def render_item(r: dict[str, Any]) -> str:
        head = (f'[{r["proposition_id"]}]\n'
                f'Proposition: {by_id[r["proposition_id"]]["text"]}')
        spans = "\n".join(
            f'Span {i}: {s.get("speaker") or "unknown"} '
            f'({s.get("speaker_side") or "unknown side"}): "{s["quote"]}"'
            for i, s in enumerate(r["spans"], 1)
        )
        return f"{head}\n{spans}"

    prompt = READJUDICATE.format(
        internal_domains=", ".join(sorted(INTERNAL_DOMAINS)),
        items="\n\n".join(render_item(r) for r in items),
    )
    text, u = call_model(prompt, max_tokens=8000)
    adj = {a["proposition_id"]: a for a in parse_json(text).get("adjudications", [])}

    rows = []
    for r in items:
        a = adj.get(r["proposition_id"], {})
        rows.append({
            "proposition_id": r["proposition_id"],
            "first_pass_verdict": r["verdict"],
            "readjudicated_verdict": a.get("verdict"),
            "reason": a.get("reason"),
            "agrees": a.get("verdict") == r["verdict"],
        })
    out = {"run_id": ex["run_id"], "usage": asdict(u), "rows": rows,
           "agreement": round(sum(1 for x in rows if x["agrees"]) / len(rows), 3)}
    (run_dir / "readjudication.json").write_text(json.dumps(out, indent=2), encoding="utf-8")
    print(json.dumps({"agreement": out["agreement"],
                      "disagreements": [x["proposition_id"] for x in rows if not x["agrees"]]},
                     indent=2))


# --------------------------------------------------------------------------
# Review sheet — the ratification protocol's paper
# --------------------------------------------------------------------------

RATIFY_CHOICES = ["Agree", "Disagree"]
SWEEP_CHOICES = ["Confirmed empty", "Missed - minor", "Missed - material"]


def review(args: argparse.Namespace) -> None:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    run_dir = Path(args.run)
    ex = json.loads((run_dir / "extraction.json").read_text(encoding="utf-8"))
    aset = json.loads(Path(args.applicable_set).read_text(encoding="utf-8"))
    by_id = {p["id"]: p for p in aset["propositions"]}

    adj: dict[str, dict[str, Any]] = {}
    adj_path = run_dir / "readjudication.json"
    if adj_path.exists():
        adj = {r["proposition_id"]: r
               for r in json.loads(adj_path.read_text(encoding="utf-8"))["rows"]}

    wb = openpyxl.Workbook()
    head_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    your_fill = PatternFill("solid", fgColor="FFF2CC")

    def style_headers(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for c in ws[1]:
            c.font = head_font
        ws.freeze_panes = "A2"

    # --- Sheet 1: committed verdicts -------------------------------------
    ws = wb.active
    ws.title = "Committed"
    ws.append(["ID", "Short", "Proposition", "Route", "Standing answer in org",
               "Machine verdict", "Confidence", "Evidence (speaker: span)",
               "Blind read", "Blind reason", "YOUR CALL", "YOUR REASON"])
    committed = [r for r in ex["records"] if r["state"] == "cited"]
    for r in committed:
        p = by_id[r["proposition_id"]]
        ev = "\n".join(
            f'{s.get("speaker") or "?"} ({s.get("speaker_side") or "?"}): "{s["quote"]}"'
            for s in r["spans"]
        )
        a = adj.get(r["proposition_id"], {})
        ws.append([r["proposition_id"], p.get("short"), p["text"], r["route"],
                   r.get("existing_human_answer"), r["verdict"], r.get("confidence"),
                   ev, a.get("readjudicated_verdict"), a.get("reason"), None, None])
    style_headers(ws, [8, 16, 40, 6, 10, 10, 10, 70, 12, 30, 12, 40])
    dv = DataValidation(type="list", formula1=f'"{",".join(RATIFY_CHOICES)}"',
                        allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    for row in range(2, len(committed) + 2):
        dv.add(ws.cell(row=row, column=11))
        ws.cell(row=row, column=11).fill = your_fill
        ws.cell(row=row, column=12).fill = your_fill
        for col in (3, 8, 10, 12):
            ws.cell(row=row, column=col).alignment = wrap

    # --- Sheet 2: unanswered — the miss sweep -----------------------------
    ws2 = wb.create_sheet("Unanswered")
    ws2.append(["ID", "Short", "Proposition", "Route", "Standing answer in org",
                "Machine state", "Machine's reason", "YOUR CALL", "WHERE / WHAT (if missed)"])
    unanswered = [r for r in ex["records"] if r["state"] != "cited"]
    for r in unanswered:
        p = by_id[r["proposition_id"]]
        reason = r.get("abstain_reason") if r["state"] == "abstain" else r.get("demoted")
        ws2.append([r["proposition_id"], p.get("short"), p["text"], r["route"],
                    r.get("existing_human_answer"), r["state"], reason, None, None])
    style_headers(ws2, [8, 16, 50, 6, 10, 10, 22, 20, 50])
    dv2 = DataValidation(type="list", formula1=f'"{",".join(SWEEP_CHOICES)}"',
                         allow_blank=True, showErrorMessage=True)
    ws2.add_data_validation(dv2)
    for row in range(2, len(unanswered) + 2):
        dv2.add(ws2.cell(row=row, column=8))
        ws2.cell(row=row, column=8).fill = your_fill
        ws2.cell(row=row, column=9).fill = your_fill
        for col in (3, 9):
            ws2.cell(row=row, column=col).alignment = wrap

    # --- Sheet 3: protocol -------------------------------------------------
    ws3 = wb.create_sheet("Protocol")
    for line in [
        "Gate 1 ratification review — ruled 29 July 2026.",
        "",
        "Committed sheet: for every machine verdict, mark Agree or Disagree.",
        "A Disagree needs a reason — reasons feed the charter.",
        "The Blind read column is a second model shown only the claim and its spans.",
        "",
        "Unanswered sheet: for every proposition the machine did not answer,",
        "mark Confirmed empty (nothing was there), Missed - minor, or",
        "Missed - material (evidence a competent seller would have flagged).",
        "For a miss, say roughly where or what — no need to find the exact quote.",
        "",
        "The bar, per transcript:",
        f"  fabricated / turn-crossing / misattributed spans: {GATE1_BAR['max_fabricated_or_crossing_spans']} allowed (mechanical)",
        f"  ratification rate: >= {GATE1_BAR['min_ratification_rate']:.0%}",
        f"  material misses: {GATE1_BAR['max_material_misses']} allowed",
        "",
        "Every cell must be filled before grading. Blanks block the grade.",
        f"Run: {ex['run_id']}   Model: {ex['model']}   Transcript sha256: {ex['transcript_sha256'][:16]}…",
    ]:
        ws3.append([line])
    ws3.column_dimensions["A"].width = 100

    out_path = run_dir / "review_sheet.xlsx"
    wb.save(out_path)
    print(json.dumps({"review_sheet": str(out_path),
                      "committed_rows": len(committed),
                      "unanswered_rows": len(unanswered)}, indent=2))


# --------------------------------------------------------------------------
# Grading the completed review against the bar
# --------------------------------------------------------------------------

def grade(args: argparse.Namespace) -> None:
    import openpyxl
    run_dir = Path(args.run)
    ex = json.loads((run_dir / "extraction.json").read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(run_dir / "review_sheet.xlsx", data_only=True)

    # mechanical: fabrication class, straight from extraction
    fabricated = sum(1 for r in ex["records"] if r["state"] == "rejected")

    committed_rows, blanks_c = [], []
    for row in wb["Committed"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        pid, call, reason = row[0], row[10], row[11]
        call = str(call).strip().lower() if call else None
        if call not in ("agree", "disagree"):
            blanks_c.append(pid)
            continue
        if call == "disagree" and not (reason and str(reason).strip()):
            blanks_c.append(f"{pid} (disagree without a reason)")
            continue
        committed_rows.append({"proposition_id": pid, "call": call,
                               "reason": (str(reason).strip() if reason else None)})

    sweep_rows, blanks_u = [], []
    for row in wb["Unanswered"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        pid, call, where = row[0], row[7], row[8]
        call = str(call).strip().lower() if call else None
        if call not in [c.lower() for c in SWEEP_CHOICES]:
            blanks_u.append(pid)
            continue
        if call.startswith("missed") and not (where and str(where).strip()):
            blanks_u.append(f"{pid} (missed without a pointer)")
            continue
        sweep_rows.append({"proposition_id": pid, "call": call,
                           "where": (str(where).strip() if where else None)})

    n = len(committed_rows)
    agreed = sum(1 for r in committed_rows if r["call"] == "agree")
    rate = round(agreed / n, 3) if n else None
    material = sum(1 for r in sweep_rows if r["call"] == "missed - material")
    minor = sum(1 for r in sweep_rows if r["call"] == "missed - minor")

    incomplete = bool(blanks_c or blanks_u)
    passed = (not incomplete
              and fabricated <= GATE1_BAR["max_fabricated_or_crossing_spans"]
              and rate is not None
              and rate >= GATE1_BAR["min_ratification_rate"]
              and material <= GATE1_BAR["max_material_misses"])

    report = {
        "run_id": ex["run_id"],
        "bar": GATE1_BAR,
        "review_incomplete": incomplete,
        "unreviewed": {"committed": blanks_c, "unanswered": blanks_u},
        "fabricated_or_crossing_spans": fabricated,
        "committed": n,
        "ratified": agreed,
        "ratification_rate": rate,
        "misses": {"material": material, "minor": minor},
        "disagreements": [r for r in committed_rows if r["call"] == "disagree"],
        "material_miss_rows": [r for r in sweep_rows if r["call"] == "missed - material"],
        "TRANSCRIPT_RESULT": ("INCOMPLETE" if incomplete else
                              ("PASS" if passed else "FAIL")),
    }
    (run_dir / "gate.json").write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(json.dumps(report, indent=2))


# --------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    def common(p):
        p.add_argument("--applicable-set", default="applicable_set.json")

    e = sub.add_parser("extract"); common(e)
    e.add_argument("--transcript", required=True)
    e.add_argument("--mode", choices=["whole", "chunked"], default="whole")
    e.add_argument("--chunk-turns", type=int, default=40)
    e.add_argument("--out", default="runs")
    e.set_defaults(func=extract)

    r = sub.add_parser("readjudicate"); common(r)
    r.add_argument("--run", required=True)
    r.set_defaults(func=readjudicate)

    v = sub.add_parser("review"); common(v)
    v.add_argument("--run", required=True)
    v.set_defaults(func=review)

    g = sub.add_parser("grade"); common(g)
    g.add_argument("--run", required=True)
    g.set_defaults(func=grade)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
